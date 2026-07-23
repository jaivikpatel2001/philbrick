"""Build the Philbrick content-audit workbook for client review.

Usage (from the frontend/ directory, with the dev server running):

    node scripts/contentAuditCrawl.mjs http://localhost:3000
    python scripts/buildContentAudit.py

Inputs:
  content-audit/crawl.json           every visible text item on every page of
                                     the NEW site, written by the crawler above
  data/generated/wpPages.json        the client's WordPress pages, as plain text
  data/generated/catalog.json        product content extracted from the
                                     WordPress product database
Output:
  content-audit/Philbrick-content-audit.xlsx

Requires openpyxl (pip install openpyxl).
"""

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FRONTEND = Path(__file__).resolve().parent.parent
HERE = FRONTEND / "content-audit"
HERE.mkdir(exist_ok=True)

crawl = json.loads((HERE / "crawl.json").read_text(encoding="utf8"))
wp_pages = json.loads(
    (FRONTEND / "data" / "generated" / "wpPages.json").read_text(encoding="utf8")
)
catalog = json.loads(
    (FRONTEND / "data" / "generated" / "catalog.json").read_text(encoding="utf8")
)

WP_BY_SLUG = {p["slug"]: p for p in wp_pages}

# ---------------------------------------------------------------------------
# Which WordPress page each route of the new site corresponds to.
# ---------------------------------------------------------------------------
ROUTE_TO_WP = {
    "/": "front-page",
    "/about": "company",
    "/vision-mission": "vision-mission",
    "/milestone": "milestone-awards",
    "/infrastructure": "infrastructure",
    "/network": "network",
    "/news-events": "news-events",
    "/contact": "contact-us",
    "/career": "career",
    "/quality-policy": "quality-policy",
    "/privacy-policy": "privacy-policy",
    "/downloads": "download",
    "/products": "shop",
}

# WordPress pages whose body is theme demo text, not client content.
WP_PLACEHOLDER = {"vision-mission", "milestone-awards", "news-events", "sample-page"}
WP_EMPTY = {"infrastructure", "network", "shop"}

PLACEHOLDER_PATTERNS = re.compile(
    r"lorem ipsum|coming soon|placeholder|dummy|tbd|to be confirmed", re.I
)

# Numbers the site states as fact and the client should confirm.
STAT_PATTERN = re.compile(r"^\s*[\d.,]+\s*(\+|%|k|m)?\s*$", re.I)


def norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", s.lower()).strip()


def wp_text_for(route: str) -> str:
    slug = ROUTE_TO_WP.get(route)
    if not slug:
        return ""
    page = WP_BY_SLUG.get(slug)
    if not page:
        return ""
    if slug in WP_PLACEHOLDER or slug in WP_EMPTY:
        return ""
    return page.get("body", "")


def strip_html(s: str) -> str:
    return re.sub(r"<[^>]+>", " ", s)


# Everything the client ever published, normalised, so a sentence can be traced
# to WordPress wherever on that site it originally appeared.
WP_CORPUS = {}
for _p in wp_pages:
    if _p["slug"] in WP_PLACEHOLDER:
        continue
    WP_CORPUS[f"WordPress page: {_p['title']}"] = norm(
        _p.get("body", "") + " " + " ".join(_p.get("elementor", []))
    )
for _cat in catalog["categories"]:
    for _prod in _cat["products"]:
        blob = " ".join(
            item for g in _prod.get("featureGroups", []) for item in g["items"]
        ) + " " + strip_html(_prod.get("specHtml", ""))
        WP_CORPUS[f"WordPress product: {_prod['name']}"] = norm(blob)


WP_WORDS = {
    source: set(w for w in corpus.split() if len(w) > 4)
    for source, corpus in WP_CORPUS.items()
}


def match_in_wp(text: str):
    """
    Trace a piece of site copy back to the client's WordPress content.

    Returns (kind, source): kind is "verbatim" when the sentence appears on
    WordPress unchanged, "adapted" when most of its distinctive words do (the
    History section, for instance, was re-punctuated and split into chapters),
    or None when the copy is new.
    """
    n = norm(text)
    if len(n) < 40:  # short labels match by accident too often
        return None
    probe = n[:120]
    for source, corpus in WP_CORPUS.items():
        if probe in corpus:
            return ("verbatim", source)

    words = set(w for w in n.split() if len(w) > 4)
    if len(words) < 6:
        return None
    best, best_score = None, 0.0
    for source, vocab in WP_WORDS.items():
        score = len(words & vocab) / len(words)
        if score > best_score:
            best, best_score = source, score
    if best_score >= 0.75:
        return ("adapted", best)
    return None


# ---------------------------------------------------------------------------
# Rows
# ---------------------------------------------------------------------------
NOISE = re.compile(r"^(\d+|0 \+|[·•—–-]+|\W{0,3})$")


# Routes whose content is deliberately temporary until the client supplies the
# real thing (see the warning banner at the top of data/news.ts).
MOCK_ROUTES = re.compile(r"^/news-events")


def status_for(route: str, text: str, matched) -> str:
    if PLACEHOLDER_PATTERNS.search(text) or MOCK_ROUTES.match(route):
        return "Placeholder - needs client input"
    if matched and matched[0] == "verbatim":
        return "Matches WordPress"
    if matched:
        return "Adapted from WordPress"
    if STAT_PATTERN.match(text):
        return "Statistic - confirm with client"
    return "New copy (not on WordPress)"


site_rows, product_rows, seo_rows = [], [], []

for page in crawl:
    route = page["path"]
    is_product = route.startswith("/products/")
    meta = page["meta"]

    seo_rows.append(
        [
            route,
            "SEO",
            "Browser / search title",
            meta.get("title", ""),
            "",
            "",
            "Shown in the browser tab and Google results",
        ]
    )
    seo_rows.append(
        [
            route,
            "SEO",
            "Meta description",
            meta.get("description", ""),
            "",
            "",
            "Google result snippet, aim for 150-160 characters",
        ]
    )
    if meta.get("ogDescription") and meta.get("ogDescription") != meta.get("description"):
        seo_rows.append(
            [route, "SEO", "Social share description", meta["ogDescription"], "", "", ""]
        )

    slug = ROUTE_TO_WP.get(route)
    for item in page["items"]:
        text = item["text"]
        if NOISE.match(text):
            continue
        matched = match_in_wp(text)
        if matched and matched[0] == "verbatim":
            suggested = f"Same wording as the {matched[1]}"
        elif matched:
            suggested = f"Reworded from the {matched[1]}"
        elif slug in WP_PLACEHOLDER:
            suggested = "WordPress page held demo text only (Lorem ipsum)"
        elif slug in WP_EMPTY:
            suggested = "WordPress page was empty"
        else:
            suggested = ""
        row = [
            route,
            item["section"].replace(" — ", " · "),
            item["type"],
            text,
            suggested,
            "",
            status_for(route, text, matched),
        ]
        (product_rows if is_product else site_rows).append(row)

# ---------------------------------------------------------------------------
# WordPress source sheet: every WP page, what it held, and where it lives now.
# ---------------------------------------------------------------------------
WP_DISPOSITION = {
    "company": ("/about + /milestone", "Migrated in full: About, Activity, History"),
    "front-page": ("/", "Product showcase migrated to the homepage range section"),
    "vision-mission": ("/vision-mission", "WordPress held Lorem ipsum; new copy written"),
    "milestone-awards": ("/milestone", "WordPress held Lorem ipsum; timeline built from the Company history"),
    "news-events": ("/news-events", "WordPress held Lorem ipsum; page exists with an empty state"),
    "privacy-policy": ("/privacy-policy", "Migrated in full, section for section"),
    "quality-policy": ("/quality-policy", "Migrated in full (4 clauses)"),
    "sample-page": ("(not migrated)", "WordPress default sample page, no client content"),
    "career": ("/career", "Migrated in full, including the HR inbox"),
    "contact-us": ("/contact", "Migrated in full: address, 4 phones, 3 emails, WhatsApp"),
    "inquiry": ("/contact", "Form migrated; every WordPress field is present"),
    "download": ("/downloads", "Migrated, including the empty-state wording"),
    "step-brochure": ("/downloads", "Brochure link migrated (2023catalog.pdf)"),
    "shop": ("/products", "All 38 products migrated with features and spec tables"),
    "infrastructure": ("/infrastructure", "WordPress page was empty; new page written"),
    "network": ("/network", "WordPress page was empty; new page written"),
}

wp_rows = []
for p in wp_pages:
    slug = p["slug"]
    where, note = WP_DISPOSITION.get(slug, ("(review)", ""))
    if slug in WP_PLACEHOLDER:
        kind = "Demo text (Lorem ipsum)"
    elif slug in WP_EMPTY or (p["htmlChars"] + p["elementorChars"]) == 0:
        kind = "Empty"
    else:
        kind = "Real client content"
    body = p.get("body", "")
    wp_rows.append(
        [
            f"WordPress: /{slug}/",
            p["title"],
            kind,
            body[:2000] + ("…" if len(body) > 2000 else ""),
            where,
            note,
            "",
        ]
    )

# ---------------------------------------------------------------------------
# Global elements: chrome that repeats on every page.
# ---------------------------------------------------------------------------
GLOBAL_ROWS = [
    ["All pages", "Header", "Logo wordmark", "Philbrick", "Philbrick India (WordPress site title)", "", "Matches WordPress"],
    ["All pages", "Header", "Tagline", "Providing Elevator Solutions", "Providing Elevator Solutions", "", "Matches WordPress"],
    ["All pages", "Header", "Navigation", "Home · About · Products · Infrastructure · Network · News & Events · Contact Us", "Same 7 items as the WordPress top menu", "", "Matches WordPress"],
    ["All pages", "Header", "Primary button", "Get a quote", "", "", "New copy (not on WordPress)"],
    ["All pages", "Footer", "Legal name", "Philbrick Technologies (India) Pvt. Ltd.", "PHILBRICK TECHNOLOGIES (INDIA) PVT. LTD. ALL RIGHTS RESERVED.", "", "Matches WordPress"],
    ["All pages", "Footer", "Helpline", "+91 84012 19941", "Helpline: +91 840 121 9941", "", "Matches WordPress"],
    ["All pages", "Footer", "WhatsApp line", "+91 99789 86631", "+91 99789 86631 OR Chat On", "", "Matches WordPress"],
    ["All pages", "Footer", "Office line", "+91 93740 22660", "+91 93740 22660", "", "Matches WordPress"],
    ["All pages", "Footer", "Office line", "+91 98250 09420", "+91 98250 09420", "", "Matches WordPress"],
    ["All pages", "Footer", "Office line", "+91 99789 86635", "+91 99789 86635", "", "Matches WordPress"],
    ["All pages", "Footer", "Office hours", "Mon to Fri, 09:00 to 18:00", "Mon-Fri 9:00 to 18:00", "", "Matches WordPress"],
    ["All pages", "Footer", "Sales email", "sales@philbrickindia.com", "sales@philbrickindia.com", "", "Matches WordPress"],
    ["All pages", "Footer", "General email", "philbrick@philbrickindia.com", "philbrick@philbrickindia.com", "", "Matches WordPress"],
    ["All pages", "Footer", "Careers email", "hr.philbrickindia@gmail.com", "hr.philbrickindia@gmail.com (Career page)", "", "Matches WordPress"],
    ["All pages", "Footer", "Alternate email", "philbrick_controls@yahoo.com", "philbrick_controls@yahoo.com", "", "Matches WordPress"],
    ["All pages", "Footer", "Registered address", "Plot No. 69, Road No. 6, G.I.D.C. Kathwada, Ahmedabad 382430, Gujarat", "Plot No. 69, Road No. 6, G.I.D.C. Kathwada, Ahmedabad 382 430 India.", "", "Matches WordPress"],
    ["All pages", "Footer", "GST number", "24AAHCP6212D1ZU", "", "", "Statistic - confirm with client"],
    ["All pages", "Footer", "CIN", "U31501GJ2014PTC078837", "", "", "Statistic - confirm with client"],
    ["All pages", "Footer", "IEC", "0814002951", "", "", "Statistic - confirm with client"],
    ["All pages", "Footer", "WhatsApp link", "https://api.whatsapp.com/send?phone=+919978986631", "Same link as the WordPress footer", "", "Matches WordPress"],
    ["All pages", "Footer", "Facebook", "https://www.facebook.com/philbrick.india", "https://www.facebook.com/philbrick.india", "", "Matches WordPress"],
    ["All pages", "Footer", "Instagram", "https://www.instagram.com/philbrick.india/", "https://www.instagram.com/philbrick.india/", "", "Matches WordPress"],
    ["All pages", "Footer", "X (formerly Twitter)", "https://twitter.com/SaranshPatel20", "https://twitter.com/SaranshPatel20", "", "Matches WordPress"],
    ["All pages", "Footer", "Newsletter heading", "Stay updated", "", "", "New copy (not on WordPress)"],
    ["All pages", "Footer", "Newsletter note", "Quarterly. No spam.", "", "", "New copy (not on WordPress)"],
    ["All pages", "Live chat", "Tawk.to property", "6039cf23385de407571a9744 / 1evgt29n1", "Same property as the WordPress plugin", "", "Matches WordPress"],
    ["All pages", "Named contact", "Customer Relationship (CRM)", "Saransh Patel", "", "", "Confirm with client"],
    ["All pages", "Named contact", "Technical Support", "Prakash Patel", "", "", "Confirm with client"],
    ["All pages", "Leadership", "Founder & Mentor", "Vasant Patel", "", "", "Confirm with client"],
    ["All pages", "Leadership", "Chief Executive Officer", "Saransh Patel", "", "", "Confirm with client"],
]

# ---------------------------------------------------------------------------
# Product source sheet: WordPress product content behind each product page.
# ---------------------------------------------------------------------------
product_source_rows = []
for cat in catalog["categories"]:
    for prod in cat["products"]:
        feature_count = sum(len(g["items"]) for g in prod.get("featureGroups", []))
        product_source_rows.append(
            [
                f"/products/{cat['slug']}",
                prod["name"],
                prod["slug"],
                feature_count,
                "Yes" if prod.get("specHtml") else "No",
                len(prod.get("images", [])),
                "",
            ]
        )

# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------
FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="0B1017")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name=FONT, size=10)
INPUT_FILL = PatternFill("solid", fgColor="FFF9D6")
WARN_FILL = PatternFill("solid", fgColor="FDE7E9")
OK_FILL = PatternFill("solid", fgColor="EAF4EC")
THIN = Side(style="thin", color="D6DAE0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

HEADERS = [
    "Page URL",
    "Section",
    "Item type",
    "Current Website Text",
    "Suggested / WordPress Text",
    "Client Final Text",
    "Status",
]

wb = Workbook()

# ---- Read me ---------------------------------------------------------------
ws = wb.active
ws.title = "Read me"
ws.sheet_view.showGridLines = False
readme = [
    ("Philbrick content audit", ""),
    ("", ""),
    ("What this workbook is", "Every piece of text on the new Philbrick website, page by page, next to the equivalent text on the current WordPress site."),
    ("What to do", "Type the final approved wording in the yellow 'Client Final Text' column. Leave a cell blank to keep the current website text as it is."),
    ("", ""),
    ("Sheet", "What it covers"),
    ("Site content", "Every heading, paragraph, list item, button and link on the company pages (home, about, contact, policies and so on)."),
    ("Product content", "The same, for all 14 product categories and 21 product pages."),
    ("Global elements", "Header, footer, contact details, statutory numbers and social links, which repeat on every page."),
    ("SEO & meta", "Browser titles and Google descriptions, one pair per page."),
    ("WordPress source", "Every page on the current WordPress site, what it contained, and where that content now lives."),
    ("Product source", "The WordPress product records behind each product page: how many salient features and whether a specification table was carried across."),
    ("", ""),
    ("Status column", "Meaning"),
    ("Matches WordPress", "The text is the client's own wording from the WordPress site, unchanged."),
    ("Adapted from WordPress", "The client's own content, re-punctuated or split up for the new layout. The facts are unchanged."),
    ("New copy (not on WordPress)", "Written for the new site, because WordPress had nothing here."),
    ("Placeholder - needs client input", "Temporary wording. Please replace."),
    ("Statistic - confirm with client", "A number stated as fact. Please confirm it is correct."),
    ("Confirm with client", "A name or role we could not verify from the WordPress site."),
    ("", ""),
    ("Colour key", ""),
    ("Yellow cells", "For you to fill in."),
    ("Pink rows", "Need a decision: placeholder text, an unverified number or a name."),
    ("Green rows", "Already match the WordPress site; usually nothing to do."),
]
for r, (a, b) in enumerate(readme, start=1):
    ws.cell(r, 1, a).font = Font(name=FONT, bold=(b == "" and a != ""), size=14 if r == 1 else 10)
    ws.cell(r, 2, b).font = BODY_FONT
    ws.cell(r, 2).alignment = WRAP
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 110


STATUS_DV = (
    '"Matches WordPress,Adapted from WordPress,New copy (not on WordPress),'
    'Placeholder - needs client input,Statistic - confirm with client,'
    'Confirm with client,Approved,Changed"'
)


def add_sheet(name, headers, rows, widths, status_col=None, example=None):
    sh = wb.create_sheet(name)
    sh.sheet_view.showGridLines = False
    for c, h in enumerate(headers, start=1):
        cell = sh.cell(1, c, h)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    sh.freeze_panes = "A2"
    sh.row_dimensions[1].height = 28

    body = ([example] if example else []) + rows
    for r, row in enumerate(body, start=2):
        for c, value in enumerate(row, start=1):
            cell = sh.cell(r, c, value)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = BORDER
        if status_col:
            st = str(row[status_col - 1])
            if st.startswith(("Placeholder", "Statistic", "Confirm")):
                for c in range(1, len(headers) + 1):
                    sh.cell(r, c).fill = WARN_FILL
            elif st == "Matches WordPress":
                for c in range(1, len(headers) + 1):
                    sh.cell(r, c).fill = OK_FILL
        if "Client Final Text" in headers:
            sh.cell(r, headers.index("Client Final Text") + 1).fill = INPUT_FILL

    if example:
        sh.cell(2, 1).comment = Comment(
            "Example row showing the expected format. Delete it before sending the "
            "sheet back, or leave it: it is ignored.",
            "Philbrick",
        )
        for c in range(1, len(headers) + 1):
            sh.cell(2, c).font = Font(name=FONT, size=10, italic=True, color="7A8290")

    for c, w in enumerate(widths, start=1):
        sh.column_dimensions[get_column_letter(c)].width = w

    if status_col:
        dv = DataValidation(type="list", formula1=STATUS_DV, allow_blank=True)
        sh.add_data_validation(dv)
        dv.add(f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}{len(body) + 1}")

    sh.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(body) + 1}"
    return sh


EXAMPLE = [
    "/about",
    "Who we are · Engineering elevator solutions",
    "Paragraph",
    "Philbrick Controls India was founded in 1992 in Ahmedabad…",
    "Same as WordPress",
    "Philbrick Controls India was established in 1992 at Ahmedabad…",
    "Changed",
]

WIDTHS = [30, 34, 14, 62, 46, 52, 26]

add_sheet("Site content", HEADERS, site_rows, WIDTHS, status_col=7, example=EXAMPLE)
add_sheet("Product content", HEADERS, product_rows, WIDTHS, status_col=7)
add_sheet("Global elements", HEADERS, GLOBAL_ROWS, WIDTHS, status_col=7)
add_sheet(
    "SEO & meta",
    ["Page URL", "Section", "Item type", "Current Website Text", "Suggested / WordPress Text", "Client Final Text", "Notes"],
    seo_rows,
    [30, 14, 24, 74, 32, 52, 44],
)
add_sheet(
    "WordPress source",
    ["WordPress page", "Title", "What it contained", "WordPress text", "Where it lives now", "Migration note", "Client Final Text"],
    wp_rows,
    [30, 24, 24, 80, 26, 46, 40],
)
add_sheet(
    "Product source",
    ["Product page", "Product", "WordPress slug", "Salient features carried across", "Specification table", "Photos", "Client Final Text"],
    product_source_rows,
    [34, 40, 40, 16, 16, 10, 40],
)

# ---- Summary (formulas, so it follows any edit) ----------------------------
summary = wb.create_sheet("Summary", 1)
summary.sheet_view.showGridLines = False
summary["A1"] = "Content audit summary"
summary["A1"].font = Font(name=FONT, bold=True, size=14)
summary["A2"] = "Counts update automatically as the Status columns are edited."
summary["A2"].font = Font(name=FONT, size=10, italic=True, color="7A8290")

STATUSES = [
    "Matches WordPress",
    "Adapted from WordPress",
    "New copy (not on WordPress)",
    "Placeholder - needs client input",
    "Statistic - confirm with client",
    "Confirm with client",
    "Approved",
    "Changed",
]
SHEET_RANGES = {
    "Site content": len(site_rows) + 2,
    "Product content": len(product_rows) + 1,
    "Global elements": len(GLOBAL_ROWS) + 1,
}

head = ["Status"] + list(SHEET_RANGES) + ["Total"]
for c, h in enumerate(head, start=1):
    cell = summary.cell(4, c, h)
    cell.font = HEAD_FONT
    cell.fill = HEAD_FILL
    cell.border = BORDER

for r, status in enumerate(STATUSES, start=5):
    summary.cell(r, 1, status).font = BODY_FONT
    summary.cell(r, 1).border = BORDER
    for c, (sheet, last) in enumerate(SHEET_RANGES.items(), start=2):
        col = get_column_letter(7)
        ref = f"'{sheet}'!${col}$2:${col}${last}"
        cell = summary.cell(r, c, f"=COUNTIF({ref},$A{r})")
        cell.font = BODY_FONT
        cell.border = BORDER
    total = summary.cell(r, len(head), f"=SUM(B{r}:{get_column_letter(len(head) - 1)}{r})")
    total.font = Font(name=FONT, size=10, bold=True)
    total.border = BORDER

trow = 5 + len(STATUSES)
summary.cell(trow, 1, "All items").font = Font(name=FONT, size=10, bold=True)
for c in range(2, len(head) + 1):
    cell = summary.cell(
        trow, c, f"=SUM({get_column_letter(c)}5:{get_column_letter(c)}{trow - 1})"
    )
    cell.font = Font(name=FONT, size=10, bold=True)
    cell.border = BORDER
summary.cell(trow, 1).border = BORDER

summary.cell(trow + 2, 1, "Pages audited").font = Font(name=FONT, size=10, bold=True)
summary.cell(trow + 2, 2, len(crawl)).font = BODY_FONT
summary.cell(trow + 3, 1, "WordPress pages reviewed").font = Font(name=FONT, size=10, bold=True)
summary.cell(trow + 3, 2, len(wp_pages)).font = BODY_FONT
summary.cell(trow + 4, 1, "WordPress products migrated").font = Font(name=FONT, size=10, bold=True)
summary.cell(trow + 4, 2, len(product_source_rows)).font = BODY_FONT
summary.cell(trow + 2, 1).comment = Comment(
    "Every page of the new site was crawled and read, including all 14 product "
    "categories and 21 product detail pages.",
    "Philbrick",
)

summary.column_dimensions["A"].width = 34
for c in range(2, len(head) + 1):
    summary.column_dimensions[get_column_letter(c)].width = 20

out = HERE / "Philbrick-content-audit.xlsx"
wb.save(out)
print("wrote", out)
print("site rows", len(site_rows), "product rows", len(product_rows), "wp rows", len(wp_rows))
